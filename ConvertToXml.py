# Loads an excel spread sheet and converts it to XML

from openpyxl import load_workbook
from yattag import Doc, indent
from datetime import date
import os

def convert2xml(cleanWorkbook, po, date):
    # Loads Excel workbook
    wb = load_workbook(cleanWorkbook)

    # Create sheet object
    ws = wb.worksheets[0]
    wblist = []

    # Returns a triplet
    doc, tag, text = Doc().tagtext()

    xml_header = '<?xml version= "1.0" encoding="UTF-8"?>'
    # Commented out by request of Wesco
   ## xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

    # Appends the string to the document
    doc.asis(xml_header)
    # Commented out by request of Wesco
   ## doc.asis(xml_schema)

    with tag('PO'):
        # Iterating rows for getting the values of each row
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=5):
            row = [cell.value for cell in row]
            wblist.append(row)
            with tag("PO_Information"):
                with tag("PO_Number"):
                    text(row[0])
                with tag("Site"):
                    text(row[1])
                with tag("PO_Description"):
                    text(row[2])
                with tag("Status"):
                    text(row[3])
                with tag("Revision"):
                    text(row[4])
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=14):
            row = [cell.value for cell in row]
            wblist.append(row)
            with tag("PO_Line"):
                with tag("Line"):
                    text(row[5])
                with tag("Line_Type"):
                    text(row[6])
                with tag("Item"):
                    text(row[7])
                with tag("Item_Description"):
                    text(row[8])
                with tag("Quantity"):
                    text(row[9])
                with tag("Order_Unit"):
                    text(row[10])
                with tag("Manufacturer"):
                    text(row[11])
                with tag("Model_Number"):
                    text(row[12])
                with tag("GL_Account"):
                    text(row[13])

    result = indent(
        doc.getvalue(),
        indentation= '  ',
        indent_text=True
    )
    ##print(result)
    with open(f"alcon_fw_{po}_{date}.xml", "w+") as f:
        f.write(result)
