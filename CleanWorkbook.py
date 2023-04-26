# Loads a spreadsheet and enters "None" in all epmty cells and saves
# the spradsheet as "cleanWorkbook.xlsx"

from openpyxl import load_workbook

## Checks for Null or "" cells
## and replace all empty cells with the string None.
def clean_WB(poWorkbook):
    wb = load_workbook(poWorkbook)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=14):
        for cell in row:
            if cell.value is None:
                cell.value = "None"
    wb.save('clean_wb.xlsx')

